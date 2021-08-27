#!/usr/bin/env python
# coding: utf-8

# In[612]:


import pdfplumber
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Color, PatternFill
from openpyxl.styles import colors, numbers
from openpyxl.styles.borders import Border, Side
import os.path
from itertools import islice
from string import ascii_uppercase
from copy import copy


# In[613]:


### FIRST PAGE ###

# page - a Page object from the pdfplumber module
# creates coordinates for address bounding box to extract text from
# returns address_bounding_box - tuple of ints/floats containing coordinates of address bounding box 
def setAddressBoundingBox(page):
    
    #get page height and width for box coordinate calculations
    page_height = page.height
    page_width = page.width
    
    #address box is near second top quarter
    top_half = page.height/2
    top_quarter = page.height/4
    second_quarter = top_half - top_quarter
    
    #coordinates of the address box
    address_box_left = (page.width//2) - 20
    address_box_top = second_quarter - 75
    address_box_right = page.width
    address_box_bottom = (page.height//2) -215
    
    address_bounding_box = (address_box_left,address_box_top,address_box_right,address_box_bottom)
    
    return address_bounding_box


# address_extract_text - str containing extracted text from address box
# returns company_name_text - str containing name of company
def setCompanyName(address_extract_text):
    
    company_name_text = address_extract_text.splitlines()[0]
    
    return company_name_text


# address_extract_text - str containing extracted text from address box
# returns street_text - str containing street name 
def setStreet(address_extract_text):

    street_text = address_extract_text.splitlines()[1]
    
    return street_text


# address_extract_text - str containing extracted text from address box
# returns city_text, state_text, zip_code_text - tuple of str containing city, state, and zip code respectively
def setCityStateZIP(address_extract_text):
    
    city_state_zip_text = address_extract_text.splitlines()[2]
    city_state_zip_text = city_state_zip_text.replace(',', '')
    
    city_text = city_state_zip_text.split(' ')[0]
    state_text =  city_state_zip_text.split(' ')[1]
    zip_code_text =  city_state_zip_text.split(' ')[2]
    
    return city_text, state_text, zip_code_text


# address_extract_text - str containing extracted text from address box
# returns acc_num_text - str containing account number
# NOT USED AS IT TURNS OUT THE ACCOUNT NUMBER IS NOT NEEDED
def setAccNum(address_extract_text): 
    
    acc_num_text = address_extract_text.splitlines()[3]
    acc_num_text = acc_num_text.split('#')[1]
    acc_num_text = acc_num_text.strip()
    
    return acc_num_text

# address_extract_text - str containing extracted text from address box
# returns issued_date_text - str containing issued date 
def setIssuedDate(address_extract_text):
    
    issued_date_text = address_extract_text.splitlines()[4]
    issued_date_text = issued_date_text.split(':')[1]
    issued_date_text = issued_date_text.strip()
    
    return issued_date_text

# page - a Page object from the pdfplumber module
# returns electric_supply_bounding_box - tuple of ints/float containing coordinates of electric supply box
def setElectricSupplyBoundingBox(page):
    
    top_half = page.height/2
    top_quarter = page.height/4
    second_quarter = top_half - top_quarter

    electric_supply_box_top  = second_quarter - 75
    electric_supply_box_bottom = (page.height//2) - 215
    electric_supply_box_left = (page.width//2) - 130
    electric_supply_box_right = (page.width//2)
    
    electric_supply_bounding_box = (electric_supply_box_left,electric_supply_box_top,electric_supply_box_right,electric_supply_box_bottom)
    
    return electric_supply_bounding_box


# page - a Page object from the pdfplumber module
# returns utility_name_text  - str containing name of utility
def setUtilityName(page):
    
    page_extract = page.extract_text(x_tolerance=1, y_tolerance=1)
    utility_name_text = page_extract.splitlines()[-6]
    
    return utility_name_text

### SECOND PAGE ###

# page - a Page object from the pdfplumber module
# creates coordinates for rate bounding box to extract text from
# returns rate_bounding_box - tuple of ints/floats containing coordinates of rate bounding box 
def setElectricRateBoundingBox(page):
    
    rate_box_left = 20
    rate_box_top = (page.width//4) + 5
    rate_box_right = (page.width//2) - 100
    rate_box_bottom = (page.width//4) + 50
    rate_bounding_box = (rate_box_left, rate_box_top, rate_box_right, rate_box_bottom)
    
    return rate_bounding_box


# rate_extract_text - str containing extracted text from rate box
# returns rate_text - str containing rate code
def setRate(rate_extract_text):

    #gets the first line after the word 'Service' and remove the leading whitespace and '-'
    rate_text = rate_extract_text.split('Service')[1].splitlines()[0].replace('-','',1).strip()
    
    if 'TOU -' in rate_text:
        
        #remove TOU and leading '-' for uniformity
        rate_text = rate_text.split('TOU')[1].replace('-','',1).strip()
    
    return rate_text

# page - a Page object from the pdfplumber module
# creates coordinates for usage bounding box to extract text from
# returns usage_bounding_box - tuple of ints/floats containing coordinates of usage bounding box 
def setElectricUsageBoundingBox(page):
    
    usage_box_left = (page.width//2) - 130
    usage_box_top = (page.height//2) - 180
    usage_box_right = (page.width//2) - 20
    usage_box_bottom = (page.height//2) - 150
    usage_bounding_box = (usage_box_left,usage_box_top,usage_box_right,usage_box_bottom)
    
    return usage_bounding_box

# page - a Page object from the pdfplumber module
# usage_bounding_box - tuple of ints/floats containing coordinates of usage bounding box
# returns usage_extract_text - str containing extracted text from usage box

def setUsage(usage_extract_text):
    
    usage_text = usage_extract_text.splitlines()[0]
    
    return usage_text


    


# In[614]:


# page - a Page object from the pdfplumber module
# returns gas_supply_bounding_box - tuple of ints/float containing coordinates of electric supply box
def setGasSupplyLowerBoundingBox(page):
    
    top_half = page.height/2
    top_quarter = page.height/4
    second_quarter = top_half - top_quarter

    gas_supply_box_top  = second_quarter + 80
    gas_supply_box_bottom = (page.height//2) + 20
    gas_supply_box_left = (page.width//2) - 130
    gas_supply_box_right = (page.width//2)
    
    gas_supply_bounding_box = (gas_supply_box_left,gas_supply_box_top,gas_supply_box_right,gas_supply_box_bottom)
    
    return gas_supply_bounding_box


# page - a Page object from the pdfplumber module
# creates coordinates for usage bounding box to extract text from
# returns usage_bounding_box - tuple of ints/floats containing coordinates of usage bounding box 
def setGasUsageRightBoundingBox(page):
    
    usage_box_left = page.width - 90
    usage_box_top = (page.height//2) - 180
    usage_box_right = page.width - 20
    usage_box_bottom = (page.height//2) - 145
    usage_bounding_box = (usage_box_left,usage_box_top,usage_box_right,usage_box_bottom)
    
    return usage_bounding_box


# In[615]:


def setGasRateRightBoundingBox(page):
    
    rate_box_left = (page.width//2) + 20
    rate_box_top = (page.width//4) + 5
    rate_box_right = (page.width) - 100
    rate_box_bottom = (page.width//4) + 50
    rate_bounding_box = (rate_box_left, rate_box_top, rate_box_right, rate_box_bottom)
    
    return rate_bounding_box


# In[616]:


#for no supplier only
def setGasOnlyUsage(usage_extract_text):
       
    usage_text = usage_extract_text.splitlines()[1]
    
    return usage_text 


# In[617]:


def setGasRateNoSuppBoundingBox(page):
    
    rate_box_left = (page.width//2) - 300
    rate_box_top = 60
    rate_box_right = (page.width//2) - 30
    rate_box_bottom = 90
    rate_bounding_box = (rate_box_left,rate_box_top,rate_box_right,rate_box_bottom)
    
    return rate_bounding_box


# In[618]:


def setGasUsageNoSuppBoundingBox(page):
    usage_box_left = (page.width//2) - 60
    usage_box_top = 100
    usage_box_right = (page.width//2) - 5
    usage_box_bottom = 130
    usage_bounding_box = (usage_box_left,usage_box_top,usage_box_right,usage_box_bottom)
    
    return usage_bounding_box


# In[619]:


def setGasRateSuppBoundingBox(page):
    rate_box_left = (page.width//2) - 300
    rate_box_top = 160
    rate_box_right = (page.width//2) - 30
    rate_box_bottom = 210
    rate_bounding_box = (rate_box_left,rate_box_top,rate_box_right,rate_box_bottom)
    
    return rate_bounding_box


# In[620]:


def setGasUsageSuppBoundingBox(page):
    
    usage_box_left = (page.width//2) - 80
    usage_box_top = 220
    usage_box_right = (page.width//2)
    usage_box_bottom = 250
    usage_bounding_box = (usage_box_left,usage_box_top,usage_box_right,usage_box_bottom)
    
    return usage_bounding_box


# In[621]:


### COMBINING DUPLICATE FUNCTIONS FOR GAS AND ELECTRIC HERE ###

def defaultExtractText(page, bounding_box):
    
    extracted_text = page.crop(bounding_box).extract_text(x_tolerance=1)
    
    return extracted_text
    
def setSupplier(extracted_text):
    
    supplier_text = extracted_text.splitlines()[1]
    
    return supplier_text

def setChoiceID(extracted_text):
    
    choice_id = extracted_text.split('Choice ID:')[1].strip()
    
    return choice_id


# In[622]:


def setMultipleLeftUsageBoundingBox(page):
    
    usage_box_left = (page.width//2) - 90
    usage_box_top = (page.height//2) - 200
    usage_box_right = (page.width//2) - 20
    usage_box_bottom = (page.height//2) + 150
    usage_bounding_box = (usage_box_left,usage_box_top,usage_box_right,usage_box_bottom)
    
    return usage_bounding_box


# In[623]:


def setMultipleRightUsageBoundingBox(page):
    
    usage_box_left = (page.width) - 90
    usage_box_top = (page.height//2) - 200
    usage_box_right = (page.width) - 20
    usage_box_bottom = page.height/2 + 150
    usage_bounding_box = (usage_box_left,usage_box_top,usage_box_right,usage_box_bottom)
    
    return usage_bounding_box


# In[624]:


def splitAndAddMultipleUsageNumbers(mult_usage_extract, unit):
    
    mult_usage_ls = mult_usage_extract.split()
    usage_numbers_ls = []
    for i in range(0,(len(mult_usage_ls))):
        if unit in mult_usage_ls[i]:
            usage_numbers_ls.append(int(mult_usage_ls[i - 1]))
            
    usage_sum = 0
    for i in range(0,len(usage_numbers_ls)):
        usage_sum += usage_numbers_ls[i]
    
    return usage_sum


# In[631]:


# sets bounding box for annual usage chart on top left of second page
# page - page from PDF 
# returns usage_bounding_box - tuple containing coordinates of bounding box for chart 
def setAnnualUsageChartBoundingBox(page):
    
    usage_box_left = 20
    usage_box_top = 20
    usage_box_right = (page.width//2) - 20
    usage_box_bottom = (page.height//2) - 230
    usage_bounding_box = (usage_box_left,usage_box_top,usage_box_right,usage_box_bottom)
    
    return usage_bounding_box


# In[632]:


# copies formatting from adjacent cell
# ws - Excel worksheet
# r - row
# c - column
def copyOriginalCellFormatting(ws, r, c):
    
    prev_cell = ws.cell(row = r, column = c-1) # gets data from the cell on the left of current cell
    
    ws.cell(row = r, column = c).font = copy(prev_cell.font)
    ws.cell(row = r, column = c).border = copy(prev_cell.border)
    ws.cell(row = r, column = c).fill = copy(prev_cell.fill)
    ws.cell(row = r, column = c).alignment = copy(prev_cell.alignment)
    ws.cell(row = r, column = c).number_format = copy(prev_cell.number_format)
    


# In[633]:


# some gas bills don't have an annual usage chart which messes up the coordinates of the bounding boxes on the second page so this check if the chart is present
# page - page from PDF 
# commodity - str specifying the commodity of the chart
# returns boolean specifying if there is an annual usage chart or not
def checkForAnnualUsageChart(page, commodity):
    
    commodity = commodity.upper()
    chart_extract = defaultExtractText(page, setAnnualUsageChartBoundingBox(page))
    
    if f'ANNUAL {commodity} USAGE' in chart_extract:
        return True
    
    else:
        return False

# generates list of letters to use as Excel column indices once the Excel file goes past col Z (next col is AA)
# returns col_list - list of strings
def generateExcelColumnList():
        
    col_list = []


    j = 0
    k = 0

    for i in range(0,26):
        col_list.append(ascii_uppercase[i])


    while col_list[len(col_list) - 1] != 'ZZ': 

        if k < 26:
            curr_col = ascii_uppercase[j] + ascii_uppercase[k]
            col_list.append(curr_col)
            k +=1
        
        else:
            
            k = 0
            j += 1
            
            curr_col = ascii_uppercase[j] + ascii_uppercase[k]
            col_list.append(curr_col)

    return col_list




# In[625]:


# file_name = input("Name of PDF file: ")


#initializing variables outside scope of open file
utility_name = ''
company_name = ''
street = ''
city = ''
state = ''
zip_code = ''
issued_date = ''
electric_supplier = ''
electric_choice_id = ''
gas_supplier = ''
gas_choice_id = ''
electric_rate = ''
electric_usage = ''
gas_rate = ''
gas_usage = ''
bill_dict = {}

supplier_present = ''
electric_supplier_present = ''
gas_supplier_present = ''

multiple_usage_electricity = ''
multiple_usage_gas = ''

excel_file_name = ''

bill_type = ''

#adding logic for different cases

# print('What type of bill is this?')
# print('e - electricity only    g - gas only    eg - electricity and gas ')
# bill_type = input('bill type: ')

# while (bill_type != 'e') and (bill_type != 'g') and (bill_type != 'eg'):
#     print('Invalid choice. Try again:')
#     bill_type = input('bill type: ').lower()

    
# if (bill_type == 'e') or (bill_type =='g'):
#     print('Is there a supplier on the bill?')
#     print('yes    no')
#     supplier_present = input().lower()
#     while (supplier_present != 'yes') and (supplier_present != 'no'):
#         print('Invalid choice. Try again:')
#         supplier_present = input().lower()
        
# if bill_type == 'e':
#     print('Are there multiple usage amounts on this bill?')
#     print('yes    no')
#     multiple_usage_electricity = input().lower()
#     while (multiple_usage_electricity != 'yes') and (multiple_usage_electricity != 'no'):
#         print('Invalid choice. Try again:')
#         multiple_usage_electricity = input().lower()
        
# if bill_type == 'g':
#     print('Are there multiple usage amounts on this bill?')
#     print('yes    no')
#     multiple_usage_gas = input().lower()
#     while (multiple_usage_gas != 'yes') and (multiple_usage_gas != 'no'):
#         print('Invalid choice. Try again:')
#         multiple_usage_gas = input().lower()
        
# if (bill_type =='eg'):
#     print('Is there an electric supplier on the bill?')
#     print('yes    no')
#     electric_supplier_present = input().lower()
#     while (electric_supplier_present != 'yes') and (electric_supplier_present != 'no'):
#         print('Invalid choice. Try again:')
#         electric_supplier_present = input().lower()
    
#     print('Is there a gas supplier on the bill?')
#     print('yes    no')
#     gas_supplier_present = input().lower()
#     while (gas_supplier_present != 'yes') and (gas_supplier_present != 'no'):
#         print('Invalid choice. Try again:')
#         gas_supplier_present = input().lower()
        
        
#     print('Are there multiple usage amounts for electricity on this bill?')
#     print('yes    no')
#     multiple_usage_electricity = input().lower()
#     while (multiple_usage_electricity != 'yes') and (multiple_usage_electricity != 'no'):
#         print('Invalid choice. Try again:')
#         multiple_usage_electricity = input().lower()
        
#     print('Are there multiple usage amounts for gas on this bill?')
#     print('yes    no')
#     multiple_usage_gas = input().lower()
#     while (multiple_usage_gas != 'yes') and (multiple_usage_gas != 'no'):
#         print('Invalid choice. Try again:')
#         multiple_usage_gas = input().lower()
         
        
# print('What is the name of the Excel file you would like to output to?')
# print('Note: Names of files are case-sensitive')
# excel_file_name = input()
# excel_file_name = excel_file_name + '.xlsx'


# In[626]:

def analyzeBill(bill_file):
    if (bill_type == 'e') and (supplier_present == 'no'): 
    
        with pdfplumber.open(bill_file) as pdf:
            first_page = pdf.pages[0]
            second_page = pdf.pages[1]

            address_extract = defaultExtractText(first_page, setAddressBoundingBox(first_page))
            #extracting from address box

            street = setStreet(address_extract)
            company_name = setCompanyName(address_extract)
            city, state, zip_code = setCityStateZIP(address_extract)
            issued_date = setIssuedDate(address_extract)

            electric_supply_extract = defaultExtractText(first_page, setElectricSupplyBoundingBox(first_page))
            #extracting from electric supply box
            electric_supplier = np.nan
            electric_choice_id = setChoiceID(electric_supply_extract) 
            
            #extracting utility name from the bottom of the page
            utility_name = setUtilityName(first_page)

            electric_rate_extract = defaultExtractText(second_page, setElectricRateBoundingBox(second_page))
            #extracting from electric_rate box
            electric_rate = setRate(electric_rate_extract)

            if multiple_usage_electricity == 'yes':
            #extracts all usage numbers and adds them
                
                multiple_usage_extract = defaultExtractText(second_page, setMultipleLeftUsageBoundingBox(second_page))
                electric_usage = splitAndAddMultipleUsageNumbers(multiple_usage_extract, 'kWh')
            
            else:
                #extracting from electric_usage
                electric_usage_extract = defaultExtractText(second_page, setElectricUsageBoundingBox(second_page))
                electric_usage = setUsage(electric_usage_extract)
            
            gas_supplier = np.nan
            gas_choice_id = np.nan
            gas_rate = np.nan
            gas_usage = np.nan

    # making dict that will be turned to df
        bill_dict = dict(utility = utility_name, issued_date = issued_date, company = company_name, 
                        street = street, city = city, state = state, zip_code = zip_code, 
                        electric_choice_id = electric_choice_id, electric_rate_code = electric_rate, 
                        electric_supplier = electric_supplier, electric_usage = electric_usage,
                        gas_supplier = gas_supplier, gas_choice_id = gas_choice_id, gas_rate_code = gas_rate, 
                        gas_usage = gas_usage)

        #10.22.20 E, 11.18.20 E format looks slightly different so doesn't work


    # In[627]:


    if (bill_type == 'e') and (supplier_present == 'yes'): 
        with pdfplumber.open(bill_file) as pdf:
            first_page = pdf.pages[0]
            second_page = pdf.pages[1]

            #extracting from address box
            address_extract = defaultExtractText(first_page, setAddressBoundingBox(first_page))

            company_name = setCompanyName(address_extract)
            street = setStreet(address_extract)
            city, state, zip_code = setCityStateZIP(address_extract)
            issued_date = setIssuedDate(address_extract)

            #extracting from electric supply box
            electric_supply_extract = defaultExtractText(first_page, setElectricSupplyBoundingBox(first_page))

            electric_supplier = setSupplier(electric_supply_extract)
            electric_choice_id = setChoiceID(electric_supply_extract)

            #extracting utility name from the bottom of the page
            utility_name = setUtilityName(first_page)

            #extracting from rate box
            rate_extract = defaultExtractText(second_page, setElectricRateBoundingBox(second_page))
            electric_rate = setRate(rate_extract)

            #extracts all usage numbers and adds them
            if multiple_usage_electricity == 'yes':

                multiple_usage_extract = defaultExtractText(second_page, setMultipleLeftUsageBoundingBox(second_page))
                electric_usage = splitAndAddMultipleUsageNumbers(multiple_usage_extract, 'kWh')
            
            else:
                #extracting from electric_usage
                electric_usage_extract = defaultExtractText(second_page, setElectricUsageBoundingBox(second_page))
                electric_usage = setUsage(electric_usage_extract)
            
            gas_supplier = np.nan
            gas_choice_id = np.nan
            gas_rate = np.nan
            gas_usage = np.nan


        bill_dict = dict(utility = utility_name, issued_date = issued_date, company = company_name, 
                        street = street, city = city, state = state, zip_code = zip_code, 
                        electric_choice_id = electric_choice_id, electric_rate_code = electric_rate, 
                        electric_supplier = electric_supplier, electric_usage = electric_usage,
                        gas_supplier = gas_supplier, gas_choice_id = gas_choice_id, gas_rate_code = gas_rate, gas_usage = gas_usage)

        #10.22.20 E, 11.18.20 E format looks slightly different so doesn't work


    # In[628]:


    if (bill_type == 'eg') and (electric_supplier_present == 'no') and (gas_supplier_present == 'no'):
        with pdfplumber.open(bill_file) as pdf:
            first_page = pdf.pages[0]
            second_page = pdf.pages[1]
            
            #extracting from address box
            address_extract = defaultExtractText(first_page, setAddressBoundingBox(first_page))

            company_name = setCompanyName(address_extract)
            street = setStreet(address_extract)
            city, state, zip_code = setCityStateZIP(address_extract)
            issued_date = setIssuedDate(address_extract)
            
            #extracting from electric supply box
            electric_supply_extract = defaultExtractText(first_page, setElectricSupplyBoundingBox(first_page))
            electric_supplier = np.nan
            electric_choice_id = setChoiceID(electric_supply_extract) 
            
            #extracting from gas supply box
            gas_supply_extract = defaultExtractText(first_page, setGasSupplyLowerBoundingBox(first_page))
            gas_choice_id = setChoiceID(gas_supply_extract)
            gas_supplier = np.nan
            
            #extracting electric rate code from second page
            rate_extract = defaultExtractText(second_page, setElectricRateBoundingBox(second_page))
            electric_rate = setRate(rate_extract)
            
            #extracts all usage numbers and adds them
            if multiple_usage_electricity == 'yes':

                multiple_usage_extract = defaultExtractText(second_page, setMultipleLeftUsageBoundingBox(second_page))
                electric_usage = splitAndAddMultipleUsageNumbers(multiple_usage_extract, 'kWh')
            
            else:
                #extracting from electric_usage
                electric_usage_extract = defaultExtractText(second_page, setElectricUsageBoundingBox(second_page))
                electric_usage = setUsage(electric_usage_extract)
            
            #extracts all usage numbers and adds them
            if multiple_usage_gas == 'yes':

                multiple_usage_extract = defaultExtractText(second_page, setMultipleRightUsageBoundingBox(second_page))
                gas_usage = splitAndAddMultipleUsageNumbers(multiple_usage_extract, 'therms')
            
            else:
                
                #extracting from gas details on second page
                gas_usage_extract = defaultExtractText(second_page, setGasUsageRightBoundingBox(second_page))
                gas_usage = setUsage(gas_usage_extract)
            
            
            gas_rate = setRate(defaultExtractText(second_page, setGasRateRightBoundingBox(second_page)))
            
            #extracting utility name from the bottom of the page
            utility_name = setUtilityName(first_page)

                    
        bill_dict = dict(utility = utility_name, issued_date = issued_date, company = company_name, 
                        street = street, city = city, state = state, zip_code = zip_code, 
                        electric_choice_id = electric_choice_id, electric_rate_code = electric_rate, 
                        electric_supplier = electric_supplier, electric_usage = electric_usage,
                        gas_supplier = gas_supplier, gas_choice_id = gas_choice_id, gas_rate_code = gas_rate, gas_usage = gas_usage)


    # In[629]:


    if (bill_type == 'eg') and (electric_supplier_present == 'yes') and (gas_supplier_present == 'yes'):
        with pdfplumber.open(bill_file) as pdf:
            first_page = pdf.pages[0]
            second_page = pdf.pages[1]
            
            #extracting from address box
            address_extract = defaultExtractText(first_page, setAddressBoundingBox(first_page))

            company_name = setCompanyName(address_extract)
            street = setStreet(address_extract)
            city, state, zip_code = setCityStateZIP(address_extract)
            issued_date = setIssuedDate(address_extract)
            
            #extracting from electric supply box
            electric_supply_extract = defaultExtractText(first_page, setElectricSupplyBoundingBox(first_page))
            electric_supplier = setSupplier(electric_supply_extract)
            electric_choice_id = setChoiceID(electric_supply_extract)
            
            #extracting from gas supply box
            gas_supply_extract = defaultExtractText(first_page, setGasSupplyLowerBoundingBox(first_page))
            gas_supplier = setSupplier(gas_supply_extract)
            gas_choice_id = setChoiceID(gas_supply_extract)
            
            #extracting electric rate code from second page
            rate_extract = defaultExtractText(second_page, setElectricRateBoundingBox(second_page))
            electric_rate = setRate(rate_extract)
            
            #extracts all usage numbers and adds them
            if multiple_usage_electricity == 'yes':

                multiple_usage_extract = defaultExtractText(second_page, setMultipleLeftUsageBoundingBox(second_page))
                electric_usage = splitAndAddMultipleUsageNumbers(multiple_usage_extract, 'kWh')
            
            else:
                #extracting from electric_usage
                electric_usage_extract = defaultExtractText(second_page, setElectricUsageBoundingBox(second_page))
                electric_usage = setUsage(electric_usage_extract)
            
            #extracts all usage numbers and adds them
            if multiple_usage_gas == 'yes':

                multiple_usage_extract = defaultExtractText(second_page, setMultipleRightUsageBoundingBox(second_page))
                gas_usage = splitAndAddMultipleUsageNumbers(multiple_usage_extract, 'therms')
            
            else:
                
                #extracting from gas details on second page
                gas_usage_extract = defaultExtractText(second_page, setGasUsageRightBoundingBox(second_page))
                gas_usage = setUsage(gas_usage_extract)
            
            gas_rate = setRate(defaultExtractText(second_page, setGasRateRightBoundingBox(second_page)))
            
            #extracting utility name from the bottom of the page
            utility_name = setUtilityName(first_page)
            
        bill_dict = dict(utility = utility_name, issued_date = issued_date, company = company_name, 
                        street = street, city = city, state = state, zip_code = zip_code, 
                        electric_choice_id = electric_choice_id, electric_rate_code = electric_rate, 
                        electric_supplier = electric_supplier, electric_usage = electric_usage,
                        gas_supplier = gas_supplier, gas_choice_id = gas_choice_id, gas_rate_code = gas_rate, gas_usage = gas_usage)


    # In[630]:


    if (bill_type == 'eg') and (electric_supplier_present == 'yes') and (gas_supplier_present == 'no'):
        with pdfplumber.open(bill_file) as pdf:
            first_page = pdf.pages[0]
            second_page = pdf.pages[1]
            
            #extracting from address box
            address_extract = defaultExtractText(first_page, setAddressBoundingBox(first_page))

            company_name = setCompanyName(address_extract)
            street = setStreet(address_extract)
            city, state, zip_code = setCityStateZIP(address_extract)
            issued_date = setIssuedDate(address_extract)
            
            #extracting from electric supply box
            electric_supply_extract = defaultExtractText(first_page, setElectricSupplyBoundingBox(first_page))
            electric_supplier = setSupplier(electric_supply_extract)
            electric_choice_id = setChoiceID(electric_supply_extract)
            
            #extracting from gas supply box
            gas_supply_extract = defaultExtractText(first_page, setGasSupplyLowerBoundingBox(first_page))
            gas_supplier = np.nan
            gas_choice_id = setChoiceID(gas_supply_extract)
            
            #extracting electric rate code from second page
            rate_extract = defaultExtractText(second_page, setElectricRateBoundingBox(second_page))
            electric_rate = setRate(rate_extract)
            
            if multiple_usage_electricity == 'yes':

                multiple_usage_extract = defaultExtractText(second_page, setMultipleLeftUsageBoundingBox(second_page))
                electric_usage = splitAndAddMultipleUsageNumbers(multiple_usage_extract, 'kWh')
            
            else:
                #extracting from electric_usage
                electric_usage_extract = defaultExtractText(second_page, setElectricUsageBoundingBox(second_page))
                electric_usage = setUsage(electric_usage_extract)
            
            if multiple_usage_gas == 'yes':

                multiple_usage_extract = defaultExtractText(second_page, setMultipleRightUsageBoundingBox(second_page))
                gas_usage = splitAndAddMultipleUsageNumbers(multiple_usage_extract, 'therms')
            
            else:
                
                #extracting from gas details on second page
                gas_usage_extract = defaultExtractText(second_page, setGasUsageRightBoundingBox(second_page))
                gas_usage = setUsage(gas_usage_extract)
            
            gas_rate = setRate(defaultExtractText(second_page, setGasRateRightBoundingBox(second_page)))
            
            #extracting utility name from the bottom of the page
            utility_name = setUtilityName(first_page)
            
        bill_dict = dict(utility = utility_name, issued_date = issued_date, company = company_name, 
                        street = street, city = city, state = state, zip_code = zip_code, 
                        electric_choice_id = electric_choice_id, electric_rate_code = electric_rate, 
                        electric_supplier = electric_supplier, electric_usage = electric_usage,
                        gas_supplier = gas_supplier, gas_choice_id = gas_choice_id, gas_rate_code = gas_rate, gas_usage = gas_usage)



    # In[634]:


    if (bill_type == 'eg') and (electric_supplier_present == 'no') and (gas_supplier_present == 'yes'):
        with pdfplumber.open(bill_file) as pdf:
            first_page = pdf.pages[0]
            second_page = pdf.pages[1]
            
            #extracting from address box
            address_extract = defaultExtractText(first_page, setAddressBoundingBox(first_page))

            company_name = setCompanyName(address_extract)
            street = setStreet(address_extract)
            city, state, zip_code = setCityStateZIP(address_extract)
            issued_date = setIssuedDate(address_extract)
            
            #extracting from electric supply box
            electric_supply_extract = defaultExtractText(first_page, setElectricSupplyBoundingBox(first_page))
            electric_supplier = np.nan
            electric_choice_id = setChoiceID(electric_supply_extract)
            
            #extracting from gas supply box
            gas_supply_extract = defaultExtractText(first_page, setGasSupplyLowerBoundingBox(first_page))
            gas_supplier = setSupplier(gas_supply_extract)
            gas_choice_id = setChoiceID(gas_supply_extract)
            
            #extracting electric rate code from second page
            rate_extract = defaultExtractText(second_page, setElectricRateBoundingBox(second_page))
            electric_rate = setRate(rate_extract)
            
            if multiple_usage_electricity == 'yes':

                multiple_usage_extract = defaultExtractText(second_page, setMultipleLeftUsageBoundingBox(second_page))
                electric_usage = splitAndAddMultipleUsageNumbers(multiple_usage_extract, 'kWh')
            
            else:
                #extracting from electric_usage
                electric_usage_extract = defaultExtractText(second_page, setElectricUsageBoundingBox(second_page))
                electric_usage = setUsage(electric_usage_extract)
            
            if multiple_usage_gas == 'yes':

                multiple_usage_extract = defaultExtractText(second_page, setMultipleRightUsageBoundingBox(second_page))
                gas_usage = splitAndAddMultipleUsageNumbers(multiple_usage_extract, 'therms')
            
            else:
                
                #extracting from gas details on second page
                gas_usage_extract = defaultExtractText(second_page, setGasUsageRightBoundingBox(second_page))
                gas_usage = setUsage(gas_usage_extract)
            
            gas_rate = setRate(defaultExtractText(second_page, setGasRateRightBoundingBox(second_page)))
            
            #extracting utility name from the bottom of the page
            utility_name = setUtilityName(first_page)
            
        bill_dict = dict(utility = utility_name, issued_date = issued_date, company = company_name, 
                        street = street, city = city, state = state, zip_code = zip_code, 
                        electric_choice_id = electric_choice_id, electric_rate_code = electric_rate, 
                        electric_supplier = electric_supplier, electric_usage = electric_usage,
                        gas_supplier = gas_supplier, gas_choice_id = gas_choice_id, gas_rate_code = gas_rate, gas_usage = gas_usage)


    # In[635]:


    if (bill_type == 'g') and (supplier_present == 'no'):
        with pdfplumber.open(bill_file) as pdf:
            first_page = pdf.pages[0]
            second_page = pdf.pages[1]
            
            #extracting from address box
            address_extract = defaultExtractText(first_page, setAddressBoundingBox(first_page))

            company_name = setCompanyName(address_extract)
            street = setStreet(address_extract)
            city, state, zip_code = setCityStateZIP(address_extract)
            issued_date = setIssuedDate(address_extract)
            
            electric_supplier = np.nan
            electric_choice_id = np.nan
            electric_rate = np.nan
            electric_usage = np.nan
            
            #extracting from gas supply box
            gas_supply_extract = defaultExtractText(first_page, setElectricSupplyBoundingBox(first_page))
            gas_choice_id = setChoiceID(gas_supply_extract)
            gas_supplier = np.nan
            
            if checkForAnnualUsageChart(second_page, 'gas') == True:
                
                if multiple_usage_gas == 'yes':

                    multiple_usage_extract = defaultExtractText(second_page, setMultipleLeftUsageBoundingBox(second_page))
                    gas_usage = splitAndAddMultipleUsageNumbers(multiple_usage_extract, 'therms')
                    
                else: 
                    gas_usage = defaultExtractText(second_page, setGasUsageSuppBoundingBox(second_page)).splitlines()[0]
                    
                gas_rate = setRate(defaultExtractText(second_page, setGasRateSuppBoundingBox(second_page)))
                
            else:  
                
                if multiple_usage_gas == 'yes':

                    multiple_usage_extract = defaultExtractText(second_page, setMultipleLeftUsageBoundingBox(second_page))
                    gas_usage = splitAndAddMultipleUsageNumbers(multiple_usage_extract, 'therms')

                else:
                    #extracting from electric_usage
                    gas_usage_extract = defaultExtractText(second_page, setGasUsageNoSuppBoundingBox(second_page))
                    gas_usage = setGasOnlyUsage(gas_usage_extract)

                #extracting electric rate code and usage from second page
        #         gas_usage_extract = defaultExtractText(second_page, setGasUsageNoSuppBoundingBox(second_page))
        #         gas_usage = setGasOnlyUsage(gas_usage_extract)
                gas_rate = setRate(defaultExtractText(second_page, setGasRateNoSuppBoundingBox(second_page)))
            
            #extracting utility name from the bottom of the page
            utility_name = setUtilityName(first_page)

        bill_dict = dict(utility = utility_name, issued_date = issued_date, company = company_name, 
                        street = street, city = city, state = state, zip_code = zip_code, 
                        electric_choice_id = electric_choice_id, electric_rate_code = electric_rate, 
                        electric_supplier = electric_supplier, electric_usage = electric_usage,
                        gas_supplier = gas_supplier, gas_choice_id = gas_choice_id, gas_rate_code = gas_rate, gas_usage = gas_usage)
                


    # In[636]:


    if (bill_type == 'g') and (supplier_present == 'yes'):
        with pdfplumber.open(bill_file) as pdf:
            first_page = pdf.pages[0]
            second_page = pdf.pages[1]
            
            #extracting from address box
            address_extract = defaultExtractText(first_page, setAddressBoundingBox(first_page))

            company_name = setCompanyName(address_extract)
            street = setStreet(address_extract)
            city, state, zip_code = setCityStateZIP(address_extract)
            issued_date = setIssuedDate(address_extract)
            
            electric_supplier = np.nan
            electric_choice_id = np.nan
            electric_rate = np.nan
            electric_usage = np.nan
            
            # extracting from gas supply box
            
            gas_supply_extract = defaultExtractText(first_page, setElectricSupplyBoundingBox(first_page))
            gas_choice_id = setChoiceID(gas_supply_extract)
            gas_supplier = setSupplier(gas_supply_extract)
            
            gas_rate = setRate(defaultExtractText(second_page, setGasRateSuppBoundingBox(second_page)))

            if multiple_usage_gas == 'yes':

                multiple_usage_extract = defaultExtractText(second_page, setMultipleLeftUsageBoundingBox(second_page))
                gas_usage = splitAndAddMultipleUsageNumbers(multiple_usage_extract, 'therms')
            
            else:
                # extracting from electric_usage
                gas_usage = defaultExtractText(second_page, setGasUsageSuppBoundingBox(second_page)).splitlines()[0]
            
            
            # extracting utility name from the bottom of the page
            utility_name = setUtilityName(first_page)

        bill_dict = dict(utility = utility_name, issued_date = issued_date, company = company_name, 
            street = street, city = city, state = state, zip_code = zip_code, 
            electric_choice_id = electric_choice_id, electric_rate_code = electric_rate, 
            electric_supplier = electric_supplier, electric_usage = electric_usage,
            gas_supplier = gas_supplier, gas_choice_id = gas_choice_id, gas_rate_code = gas_rate, gas_usage = gas_usage)

    return bill_dict
      
def createBillDf(bill_file):
    bill_df = pd.DataFrame([analyzeBill(bill_file)])
    # bill_df = pd.DataFrame([analyzeBill(f'Bills/{file_name}.pdf')])


    # In[637]:


    # bill_df = pd.DataFrame([bill_dict])
    bill_df.columns = [col.replace('_',' ') for col in bill_df]
    bill_df.columns = [col.title() for col in bill_df]
    bill_df.rename(columns = {
        'Electric Choice Id': 'Electric Choice ID',
        'Electric Usage' : 'Electric Usage (kWh)',
        'Gas Choice Id' : 'Gas Choice ID',
        'Gas Usage' : 'Gas Usage (dth)',
        'Zip Code' : 'ZIP Code'
    }, inplace = True)

    bill_df['Electric Usage (kWh)'] = pd.to_numeric(bill_df['Electric Usage (kWh)'])
    bill_df['Electric Usage (kWh)'] = pd.to_numeric(bill_df['Electric Usage (kWh)']) * 12 # multiplying by 12 for entire year

    bill_df['Gas Usage (dth)'] = pd.to_numeric(bill_df['Gas Usage (dth)'])
    bill_df['Gas Usage (dth)'] = (bill_df['Gas Usage (dth)'] / 10) * 12 # converting to decatherms and multiplying by 12 for entire year

    bill_df['Issued Date'] = pd.to_datetime(bill_df['Issued Date'])
    bill_df['Issued Date'] = bill_df['Issued Date'].dt.strftime('%m/%d/%Y')

    bill_df.fillna('', inplace = True)


    # In[638]:


    bill_df['Electric Service State'] = bill_df['State']
    bill_df['Gas Service State'] = bill_df['State']
    bill_df['Gas Utility'] = bill_df['Utility']

    bill_df.rename(columns = {
        'Utility' : 'Electric Utility'
    }, inplace = True)

    # if bill_type == 'e':
    #     bill_df['Gas Utility'] = 'N/A'
    #     bill_df['Gas Service State'] = 'N/A'
        
    # if bill_type == 'g':
    #     bill_df['Electric Utility'] = 'N/A'
    #     bill_df['Electric Service State'] = 'N/A'


    # In[639]:


    bill_df = bill_df[['Company',
    'Issued Date',
    'Street',
    'City',
    'State',
    'ZIP Code',
    'Electric Choice ID',
    'Electric Service State',
    'Electric Utility',
    'Electric Rate Code','Electric Usage (kWh)',
    'Electric Supplier',
    'Gas Choice ID',
    'Gas Service State',
    'Gas Utility',
    'Gas Rate Code',
    'Gas Usage (dth)',
    'Gas Supplier']]

    return bill_df



# In[640]:

def writeToExcelFile(bill_df, excel_file_name):
    
    execution_complete = False

    # creating variables to hold formatting for header cells
    gray_fill = PatternFill(start_color='D0D3D6', end_color='D0D3D6', fill_type='solid')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # if file exists already, get the data from the 'Unformatted Data' sheet and append new data to the end
    if os.path.isfile(excel_file_name) == True:
        wb = openpyxl.load_workbook(excel_file_name)
        
        if 'Unformatted Data' in wb.sheetnames:
            
            ws_unformatted = wb['Unformatted Data']
            
            for r in dataframe_to_rows(bill_df, index=False, header=False):
                ws_unformatted.append(r)
                
            # delete the old formatted sheet and replace it with a new blank one since it isn't easy to append to a new columns
    #         wb.remove(wb['Formatted Data'])
        
        else: 
            ws_unformatted = wb.create_sheet('Unformatted Data')
        
            for r in dataframe_to_rows(bill_df, index=False, header=True):
                ws_unformatted.append(r)

            
        for i in range(0, 18):
            ws_unformatted.cell(row = 1, column = i+1).font = Font(bold = True)
            ws_unformatted.cell(row = 1, column = i+1).fill = gray_fill
            ws_unformatted.cell(row = 1, column = i+1).border = thin_border
        
        # since for some reason the company column never gets appended correctly, I manually extract it 
        # from the second sheet and add it back
        company_col = []

        for cell in ws_unformatted['A']:
            company_col.append(cell.value)
            
        data = ws_unformatted.values
        cols = next(data)[1:]
        data = list(data)
        data = (islice(r, 1, None) for r in data)
        existing_bill_df = pd.DataFrame(data, columns=cols)
        existing_bill_df.insert(0, column = company_col[0], value = company_col[1:])

        ws_unformatted.protection.sheet = True
        ws_unformatted.sheet_state = 'hidden'
        
        ws_transposed = wb.create_sheet('Formatted Data')
        
        # since the df is transposed, openpyxl thinks the headers in the first column is the index
        for r in dataframe_to_rows(existing_bill_df.T, index=True, header=False):
            ws_transposed.append(r)
            
        
        for i in range(1, 19):
            ws_transposed.cell(column = 1, row = i+1).font = Font(bold = True)
            ws_transposed.cell(column = 1, row = i+1).fill = gray_fill
            ws_transposed.cell(column = 1, row = i+1).border = thin_border
                
        # to insert breaks in between gas and electric
        ws_transposed.delete_rows(1)
        ws_transposed.insert_rows(3)
        ws_transposed.insert_rows(14)
        
        ws_transposed.protection.sheet = True
        ws_transposed.sheet_state = 'hidden'
        
    else:
        
        raise Exception('Excel file not found')

        # # create new workbook
        # wb = openpyxl.Workbook()
        # ws_unformatted = wb.active
        # ws_unformatted.title = 'Unformatted Data'
        
        # for r in dataframe_to_rows(bill_df, index=False, header=True):
        #     ws_unformatted.append(r)
            
        # for i in range(0, 18):
        #     ws_unformatted.cell(row = 1, column = i+1).font = Font(bold = True)
        #     ws_unformatted.cell(row = 1, column = i+1).fill = gray_fill
        #     ws_unformatted.cell(row = 1, column = i+1).border = thin_border
        
        # ws_unformatted.protection.sheet = True
        # ws_unformatted.sheet_state = 'hidden'
        
        # ws_transposed = wb.create_sheet('Formatted Data', 0)
        
        # for r in dataframe_to_rows(bill_df.T, index=True, header=False):
        #     ws_transposed.append(r)
        
        # for i in range(1, 19):
        #     ws_transposed.cell(column = 1, row = i+1).font = Font(bold = True)
        #     ws_transposed.cell(column = 1, row = i+1).fill = gray_fill
        #     ws_transposed.cell(column = 1, row = i+1).border = thin_border
        
        # ws_transposed.delete_rows(1)
        # ws_transposed.insert_rows(3)
        # ws_transposed.insert_rows(14)


    # In[641]:


    ws_flight_plan = wb['1 - Flight Plan']

    transposed_max_col = ws_transposed.max_column
        
    # ELECTRICITY

    # raw data starts in col B but flight plan starts in G
    transposed_col_offset = 7-2

    for i in range (4, 14):
        for j in range (2, transposed_max_col + 1): # start from col 2 because that's where the data is in the transposed sheet 
            
            c = ws_transposed.cell(row = i, column = j)  
            ws_flight_plan.cell(row = i, column = j + transposed_col_offset).value = c.value
            
            if j > 6: # add more headers if there are more than 5 locations
            
                ws_flight_plan.cell(row = 3, column = j + transposed_col_offset).value = f'Location {j-1}'
            
    # GAS

    # offset is needed since gas data in transposed sheet starts on row 15 but flight plan starts on row 21
    transposed_gas_row_offset = 21-15

    for i in range (15, 21):
        for j in range (2, transposed_max_col + 1):

            c = ws_transposed.cell(row = i, column = j) # get data from transposed sheet
    
            ws_flight_plan.cell(row = i + transposed_gas_row_offset, column = j + transposed_col_offset).value = c.value # put data into flight plan
        
            if j > 6: # add more headers if there are more than 5 locations
            
                ws_flight_plan.cell(row = 20, column = j + transposed_col_offset).value = f'Location {j-1}'
    
    excel_col_list = generateExcelColumnList()

    # if there are more than 5 locations, adjust Excel formulas past col E to match 
    if ws_transposed.max_column > 6:
        
        last_col = excel_col_list[ws_flight_plan.max_column-1]

        ws_flight_plan['E10'].value = f'=SUM(G12:{last_col}12)'
        ws_flight_plan['E13'].value = f'=SUM(G18: {last_col}18)'
        ws_flight_plan['E23'].value = f'=SUM(G25:{last_col}25)'
        ws_flight_plan['E26'].value = f'=SUM(G31: {last_col}31)'

        for i in range(12,ws_flight_plan.max_column + 1):

            curr_col = excel_col_list[i - 1]
            ws_flight_plan[f'{curr_col}18'] = f'= {curr_col}12 * {curr_col}16'
            ws_flight_plan[f'{curr_col}31'] = f'= {curr_col}25 * {curr_col}29'
                
    # matching formatting of new columns to original template     
    for i in range (3, 32):
        for j in range (11, ws_flight_plan.max_column + 1):
            
            copyOriginalCellFormatting(ws_flight_plan, i, j)


    # if ws_flight_plan.max_column > 6: # if there are more than 5 locations, adjust Excel formulas in col E to match
        
    #     if (ws_flight_plan.max_column > 26):

    #         curr_col_num = 27
    #         first_col_letter = 0
    #         second_col_letter = 0

    #         while (len(bill_df.index) > curr_col_num):
    #             curr_col_num +=1
    #             second_col_letter += 1

    #             if second_col_letter > 26:
    #                 first_col_letter += 1
    #                 second_col_letter = 0
        
    #         last_col = ascii_uppercase[first_col_letter] + ascii_uppercase[second_col_letter]
        
    #     else:
    #         last_col = ascii_uppercase[ws_flight_plan.max_column-1]

    #     ws_flight_plan['E10'].value = f'=SUM(G12:{last_col}12)'
    #     ws_flight_plan['E13'].value = f'=SUM(G18: {last_col}18)'
    #     ws_flight_plan['E23'].value = f'=SUM(G25:{last_col}25)'
    #     ws_flight_plan['E26'].value = f'=SUM(G31: {last_col}31)'
        
    #     first_col_letter_curr_col = 0
    #     second_col_letter_curr_col = 0

    #     # adding Excel formula at the bottom of each commodity column to get cost for each location
    #     for i in range(12,ws_flight_plan.max_column + 1):
    #         for count, col_name in enumerate(ascii_uppercase):
            
    #             if (i > 26):

    #                 curr_col = ascii_uppercase[first_col_letter_curr_col] + ascii_uppercase[second_col_letter_curr_col]

    #                 second_col_letter += 1

    #                 if second_col_letter > 26:
    #                     first_col_letter += 1
    #                     second_col_letter = 0

    #             else:
    #                 curr_col = ascii_uppercase[i - 1]
        
    #             ws_flight_plan[f'{curr_col}18'] = f'= {curr_col}12 * {curr_col}16'
    #             ws_flight_plan[f'{curr_col}31'] = f'= {curr_col}25 * {curr_col}29'

    # matching formatting of new columns to original template     
    for i in range (3, 32):
        for j in range (11, ws_flight_plan.max_column + 1):
            
            copyOriginalCellFormatting(ws_flight_plan, i, j)
        
    wb.remove(wb['Formatted Data']) # This sheet is no longer needed after the data is copied to the flight plan

    wb.save(excel_file_name)
    
    execution_complete = True

    return execution_complete

# analyzed_bill = createBillDf(file_name)
# writeToExcelFile(analyzed_bill, excel_file_name)

